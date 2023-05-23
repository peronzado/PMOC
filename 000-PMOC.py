#coding: utf-8

import openpyxl
import datetime
import numpy as np


#variaveis que precisam ser apagadas um_ano_depois
#numero do contrato
contrato = 333333


#arquivos de data para mencionar a vigência:
hoje = datetime.date.today()
um_ano_depois = hoje.replace(day=hoje.day-1,year=hoje.year+1)
hoje = hoje.strftime("%d/%m/%Y")
um_ano_depois = um_ano_depois.strftime("%d/%m/%Y")

# Carrega o arquivo do Excel
workbook = openpyxl.load_workbook('000-Inspecao.xlsx')

# Seleciona a planilha desejada
worksheet = workbook['Relatorio']

# Itera sobre as células da tabela
for row in worksheet.iter_rows(min_row=5, min_col=2, max_col=11):

    # Cria uma lista com os valores da linha atual
    row_values = [cell.value for cell in row]

    # Verifica se a linha é nula (todos os valores são None)
    if all(value is None for value in row_values):
        continue


    # Aqui você pode fazer o que quiser com a lista de valores da linha
    #print(row_values)



# Cria uma lista vazia para armazenar os valores da tabela
table_data = []

# Cria uma lista com os cabeçalhos das colunas
#headers = ['#', 'vara', 'Fabr.', 'vare', 'Nº série', 'vard', 'Oc. Fixos', 'Oc. Flut.', 'Pot. (Btu/h)', 'Pot. (TR)']
table = ''


# Itera sobre as células da tabela
for row in worksheet.iter_rows(min_row=5, min_col=2, max_col=11):
    # Cria uma lista com os valores da linha atual
    row_values = [cell.value for cell in row]

    # Verifica se a linha é nula (todos os valores são None)
    if all(value is None for value in row_values):
        continue

    # Adiciona o número da linha à lista de valores da linha
    #row_values.insert(0, len(table_data) + 1)

    # Converte os valores da linha para strings e adiciona à tabela em LaTeX
    table += ' & '.join([str(value) for value in row_values]) + r' \\ ' + '\n' + '\hline '

# Finaliza a tabela em LaTeX
#table += r'\bottomrule' + '\n'
#table += r'\end{tabular}' + '\n'
#table += r'\end{table}' + '\n'




# Cria o arquivo Main
pre_relacao = ""
pre_relacao +="\\documentclass[a4paper, 12pt]{article}\n"
pre_relacao +="\n"
pre_relacao +="\\usepackage{cmbright}\n"
pre_relacao +="\\usepackage[OT1]{fontenc}\n"
pre_relacao +="\\usepackage{adjustbox}\n"
pre_relacao +="\\usepackage{tabularx}\n"
pre_relacao +="\\usepackage{multirow}\n"
pre_relacao +="\n"
pre_relacao +="\\usepackage[dvipsnames]{xcolor}\n"
pre_relacao +="\n"
pre_relacao +="\\usepackage{enumitem}\n"
pre_relacao +="\n"
pre_relacao +="\n"
pre_relacao +="% Importing document settings from our file (packages.sty)\n"
pre_relacao +="\\usepackage{packages}\n"
pre_relacao +="\n"
pre_relacao +="\\newcommand\\BackgroundCoverPic{\n"
pre_relacao +="    \\put(0,0){\n"
pre_relacao +="    \\parbox[b][\\paperheight]{\\paperwidth}{%\n"
pre_relacao +="    \\vfill\n"
pre_relacao +="    \\centering\n"
pre_relacao +="    \\includegraphics[width=\\paperwidth,height=\\paperheight]{000-Cover.pdf}\n"
pre_relacao +="    \\vfill\n"
pre_relacao +="    }}}\n"
pre_relacao +="    \n"
pre_relacao +="\\newcommand\\BackgroundPic{\n"
pre_relacao +="    \\put(0,0){\n"
pre_relacao +="    \\parbox[b][\\paperheight]{\\paperwidth}{%\n"
pre_relacao +="    \\vfill\n"
pre_relacao +="    \\centering\n"
pre_relacao +="    \\includegraphics[width=\\paperwidth,height=\\paperheight]{000-BackGround.pdf}\n"
pre_relacao +="    \\vfill\n"
pre_relacao +="    }}}\n"
pre_relacao +="    \n"
pre_relacao +="\\newcommand\\ART{\n"
pre_relacao +="    \\put(0,0){\n"
pre_relacao +="    \\parbox[b][\\paperheight]{\\paperwidth}{%\n"
pre_relacao +="    \\vfill\n"
pre_relacao +="    \\centering\n"
pre_relacao +="    \\includegraphics[width=\\paperwidth,height=\\paperheight]{000-ART.pdf}\n"
pre_relacao +="    \\vfill\n"
pre_relacao +="    }}}\n"
pre_relacao +="\n"
pre_relacao +="% Beginning of document\n"
pre_relacao +="\\begin{document}\n"
pre_relacao +="\n"
pre_relacao +="\\AddToShipoutPicture*{\\BackgroundCoverPic}\n"
pre_relacao +="% \\maketitle\n"
pre_relacao +="\n"
pre_relacao +="% Inserting title page\n"
pre_relacao +="\\import{./}{title}\n"
pre_relacao +="\n"
pre_relacao +="% Defining front matter settings (Norsk: innstillinger for forord m.m.)\n"
pre_relacao +="\\frontmatter\n"
pre_relacao +="\n"
pre_relacao +="\\AddToShipoutPicture{\\BackgroundPic}\n"
pre_relacao +="\n"
pre_relacao +="% Inserting table of contents\n"
pre_relacao +="\\tableofcontents\n"
pre_relacao +="\n"
pre_relacao +="% Inserting list of figures & list of tables\n"
pre_relacao +="%\\listoffigures\n"
pre_relacao +="%\\listoftables\n"
pre_relacao +="\n"
pre_relacao +="% Defining main matter settings (Norsk: innstillinger for hoveddelen av teksten)\n"
pre_relacao +="\\mainmatter\n"
pre_relacao +="\n"
pre_relacao +="% Introduction explaining this LaTeX-template\n"


pre_relacao +="\\section{Introdução}\n"
pre_relacao +="Introduction section\n"
pre_relacao +="\\section{Relação de Equipamentos}\n"
pre_relacao +="\n"
pre_relacao +="\\begin{center}\n"
pre_relacao +="\\begin{adjustbox}{width=\\columnwidth,center}\n"
pre_relacao +="\\begin{tabular}{|c|c|c|c|c|c|c|c|c|c|}\n"
pre_relacao +="\\hline$\\#$ & vara & varb. & varc & N $^{\\circ}$ vard. & vare & Oc. varf & varg & varh & vari \\\\\n"
pre_relacao +="\\hline\n"




# inserir nesse espaço a tabela relacao de equipamentos

relacao ="\\end{tabular}\n"
relacao +="\\end{adjustbox}\n"
relacao +="\\end{center}\n"
relacao +="\n"
relacao +="\\begin{tabular}{|l|l|}\n"


# inserir nesse espaço a tabela com as potências instaladas
# Realizando a soma das colunas 9 e 10
# inicia as variáveis de soma
soma_x = 0
soma_y = 0

# itera pelas linhas a partir da quinta linha
for row in worksheet.iter_rows(min_row=5, min_col=10, values_only=True):
    # adiciona os valores das colunas J e K à soma correspondente

    if all(value is None for value in row):
        continue

    soma_x += row[0]
    soma_y += row[1]


# Criando a tabela no LaTeX
tabela_latex = ""
tabela_latex += "\\hline \\textbf{Somax } & " + str(soma_x) + " \\\\\n"
tabela_latex += "\\hline \\textbf{Somay } & " + str(soma_y) + " \\\\\n"


pos_relacao ="\\hline\n"
pos_relacao +="\\end{tabular}\n"
pos_relacao +="\n"
pos_relacao +="\\section{Procedimento}\n"
pos_relacao +="Introduzir o procedimento"
pos_relacao +="\\section{Lista Total}\n"

lista_total = ''
# inserir aqui o codigo para os planos individuais por maquina
# Itera pelas linhas não nulas da tabela
for row in worksheet.iter_rows(min_row=5, min_col=2, max_col=11, values_only=True):

    if all(value is None for value in row):
        continue
    # Extrai os valores de cada coluna da linha atual
    hashtag, vara, varb, varc, vard, vare, varf, varg, varh, vari = row



    lista_total +="\\subsection{vara: " + str(vara) +" - " + str(vard) + "}\n"
    lista_total +="\n"
    lista_total +="\n"
    lista_total +="\n"
    lista_total +="\\begin{adjustbox}{width=18cm ,center}\n"
    lista_total +="\\begin{tabular}{|c|c|c|c|c|c|c|c|}\n"
    lista_total +="\\hline vara: & " + str(vara) +" & varb: & " + str(varb) +" & $N^{\\circ}$ SÉRIE & " + str(varc) +" & \\textbf{Página deste Equipamento} & \\textbf{1/2} \\\\\n"
    lista_total +="\\hline vard & " + str(vard) +" & vare: & " + str(vare) +" & varf: & "+ str(hoje) + " & varg: & "+ str(um_ano_depois) + " \\\\\n"
    lista_total +="\\hline\n"
    lista_total +="\\end{tabular}\n"
    lista_total +="\\end{adjustbox}\n"
    lista_total +="\n"
    lista_total +="\\begin{adjustbox}{width=18cm ,center}\n"
    lista_total +="\\begin{tabular}{|c|p{12cm}|c|c|c|c|c|c|c|c|c|c|c|c|c|c|}\n"
    lista_total +="\\hline $\\mathrm{N}^0$ & SERVIÇOS & Frequência (dias) & month & month & month & month & month & month & month & month & month & month & month & month & month \\\\\n"
    lista_total +="\\hline 1 &  Chore & " + str(int(30/1)) +" &  &  & &  & &  &  & &  & &  &  &  \\\\\n"
    lista_total +="\\hline\n"
    lista_total +="\\hline\n"
    lista_total +="\\multicolumn{3}{|c|}{\\textbf{Data }} &  & & & & & & & &  &  & &  &  \\\\\n"
    lista_total +="\\hline \\multicolumn{3}{|c|}{\\textbf{Assinatura }} &  & & & & & & & & & &   &  &  \\\\\n"
    lista_total +="\\hline \\multicolumn{3}{|c|}{\\textbf{Assinatura }} &  & & & & & & & & & &   &  &  \\\\\n"
    lista_total +="\\hline\n"
    lista_total +="\\end{tabular}\n"
    lista_total +="\\end{adjustbox}\n"
    lista_total +="\n"
    lista_total +="\n"
    lista_total +="\\footnotesize\\textbf{Observações:}\\\\[0.1cm]\n"
    lista_total +="\\noindent\\rule{\\textwidth}{0.2mm} \\\\[0.1cm]\n"
    lista_total +="\\noindent\\rule{\\textwidth}{0.2mm} \\\\[0.1cm]\n"
    lista_total +="\\noindent\\rule{\\textwidth}{0.2mm} \\\\[0.1cm]\n"
    lista_total +="\\noindent\\rule{\\textwidth}{0.2mm} \\\\[0.1cm]\n"
    lista_total +="\\noindent\\rule{\\textwidth}{0.2mm}\n"
    lista_total +="\n"
    lista_total +="\n"
    lista_total +="\n"
    lista_total +="\n"
    lista_total +="\\newpage\n"
    lista_total +="\\begin{adjustbox}{width=18cm ,center}\n"
    lista_total +="\\begin{tabular}{|c|c|c|c|c|c|c|c|}\n"
    lista_total +="\\hline vara: & " + str(vara) +" & varb: & " + str(varb) +" & $N^{\\circ}$ SÉRIE & " + str(varc) +" & \\textbf{Página deste Equipamento} & \\textbf{2/2} \\\\\n"
    lista_total +="\\hline vard & " + str(vard) +" & vare: & " + str(vare) +" & varf: & "+ str(hoje) + " & varg: & "+ str(um_ano_depois) + " \\\\\n"
    lista_total +="\\hline\n"
    lista_total +="\\end{tabular}\n"
    lista_total +="\\end{adjustbox}\n"
    lista_total +="\n"
    lista_total +="\\begin{adjustbox}{width=18cm ,center}\n"
    lista_total +="\\begin{tabular}{|c|p{12cm}|c|c|c|c|c|c|c|c|c|c|c|c|c|c|}\n"
    lista_total +="\\hline $\\mathrm{N}^0$ & SERVIÇOS & Frequência (dias) & Nov & Dez & Jan & Fev & Mar & Abr & Mai & Jun & Jul & Ago & Set & Out & Nov \\\\\n"
    lista_total +="\\hline 18 & tarefa & " + str(int(30/1)) +" & & & & & & & & & & & & & \\\\\n"
    lista_total +="\\hline\n"
    lista_total +="\\multicolumn{3}{|c|}{\\textbf{Data }} &  & & & & & & & &  &  & &  &  \\\\\n"
    lista_total +="\\hline \\multicolumn{3}{|c|}{\\textbf{Assinatura }} &  & & & & & & & & & &   &  &  \\\\\n"
    lista_total +="\\hline \\multicolumn{3}{|c|}{\\textbf{Assinatura }} &  & & & & & & & & & &   &  &  \\\\\n"
    lista_total +="\\hline\n"
    lista_total +="\\end{tabular}\n"
    lista_total +="\\end{adjustbox}\n"
    lista_total +="\n"
    lista_total +="\n"
    lista_total +="\\footnotesize\\textbf{Observações:}\\\\[0.1cm]\n"
    lista_total +="\\noindent\\rule{\\textwidth}{0.2mm} \\\\[0.1cm]\n"
    lista_total +="\\noindent\\rule{\\textwidth}{0.2mm} \\\\[0.1cm]\n"
    lista_total +="\\noindent\\rule{\\textwidth}{0.2mm} \\\\[0.1cm]\n"
    lista_total +="\\noindent\\rule{\\textwidth}{0.2mm} \\\\[0.1cm]\n"
    lista_total +="\\noindent\\rule{\\textwidth}{0.2mm}\n"
    lista_total +="\\newpage\n"




final ="\\newpage\n"
final +="\\section{ANEXOS (Ordens de Serviço e N1 do prestador)}\n"
final += '\end{document}'

#alterar posteriormente o main para receber todos os valores incluindo as tabelas
main = pre_relacao + table + relacao + tabela_latex + pos_relacao + lista_total + final


# Cria o arquivo title
Title1 = ''
Title2 = ''
Title3 = ''
Title4 = ''
Title1 +="% Inspired by title template from ShareLaTeX Learn; Gubert Farnsworth & John Doe\n"
Title1 +="% Edited by Jon Arnt Kårstad, NTNU IMT\n"
Title1 +="\n"
Title1 +="\\begin{titlepage}\n"
Title1 +="%\\vbox{ }\n"
Title1 +="%\\vbox{ }\n"
Title1 +="\n"
Title1 +="\\begin{center}\n"
Title1 +="% Upper part of the page\n"
Title1 +="\\includegraphics[width=1\\textwidth]{000-Logo}\\\\[2cm]\n"
Title1 +="\\vbox{ }\n"
Title1 +="\\vbox{ }\n"
Title1 +="\n"
Title1 +="\\color{Sepia}\n"
Title1 +="\\textsc{\\huge \\textbf{contrato}}\\\\[2.5cm]\n"
Title1 +="\\textsc{\\LARGE contrato "
Title2 +="}\\\\[2.5cm]\n"
Title2 +="\\vbox{ }\n"
Title2 +="\n"
Title2 +="\\color{Gray}\n"
Title2 +="\n"
Title2 +="% Title\n"
Title2 +="%\\HRule \\\\[0.4cm]\n"
Title2 +="{ \\Large \\b1eries Data: "
Title3 +="}\\\\[0.25cm]\n"
Title3 +="{ \\large \\b1eries Validade: "
Title4 +="}\\\\[3.5cm]\n"
Title4 +="\\HRule \\\\[0.05cm]\n"
Title4 +="\n"
Title4 +="% Author\n"
Title4 +="\\textbf{\\emph{Nome da Empresa}}\\\\\n"
Title4 +="\\emph{CNPJ:}\n"
Title4 +="CNPJ\\\\\n"
Title4 +="\\emph{CREA-SP:}\n"
Title4 +="\n"
Title4 +="\\vfill\n"
Title4 +="\n"
Title4 +="% Bottom of the page\n"
Title4 +="%{\\large Date}\n"
Title4 +="\\end{center}\n"
Title4 +="\\end{titlepage}\n"


Title = Title1 + str(contrato) + Title2 + hoje + Title3 + um_ano_depois + Title4

# Escrita dos arquivos

# Titulo
with open('title.tex', mode="w", encoding="utf-8") as f:
    f.write(Title)

# Main
with open('main.tex', mode="w", encoding="utf-8") as g:
    g.write(main)

# Arquivo Packages.sty, importado de Jon Arnt Kårstad, NTNU IV-IMT
packages = ""
packages +="% Author - Jon Arnt Kårstad, NTNU IV-IMT\n"
packages +="\n"
packages +="\\NeedsTeXFormat{LaTeX2e}\n"
packages +="\\ProvidesPackage{packages}[2021/04/20 Packages]\n"
packages +="\\usepackage[utf8]{inputenc} % Character encoding\n"
packages +="\n"
packages +="% ------ Contents -------\n"
packages +="% Appendices\n"
packages +="% Bibliography & References\n"
packages +="% Code input\n"
packages +="% Colors\n"
packages +="% Flow charts\n"
packages +="% Front matter\n"
packages +="% General\n"
packages +="% Image\n"
packages +="% Language\n"
packages +="% Main matter\n"
packages +="% Mathematics\n"
packages +="% Page setup\n"
packages +="% -----------------------\n"
packages +="\n"
packages +="% --- Appendices ---\n"
packages +="\\usepackage{appendix}\n"
packages +="\\newcommand{\\addappendix}{   % Self-created command to insert appendix with predefined settings\n"
packages +="    \\newpage\n"
packages +="    \\appendix\n"
packages +="    \\section*{Appendix}   % Name of appendix\n"
packages +="    \\addcontentsline{toc}{section}{Appendix}  % Add appendix name to table of contents\n"
packages +="    \\renewcommand{\\thesubsection}{\\Alph{subsection}}    % Change numbering of section to upper-case letters.\n"
packages +="}\n"
packages +="\n"
packages +="% --- Bibliography & References---  \n"
packages +="\\usepackage[backend = biber,    % Recommended backend for sorting bibliography\n"
packages +="            style = authoryear-comp,    % Close to the 'Harvard' referencing style\n"
packages +="            urldate = long,     % Long: 24th Mar. 1997 | Short: 24/03/1997\n"
packages +="            maxcitenames = 2,   % Number of authors in cite before replaced with 'Author#1 et al.'\n"
packages +="            ]{biblatex}\n"
packages +="\\addbibresource{references.bib}     % Adding our file containing the references\n"
packages +="% \\addbibresource{reference_2.bib} is possible if we want to add several reference files\n"
packages +="\n"
packages +="\\usepackage{caption}        % Enables controlling the look and feel of captions, see package documentation\n"
packages +="\\usepackage{subcaption}     % Recommended when making sub-figures\n"
packages +="\\usepackage[nottoc]{tocbibind}  % Includes Bibliography, Index, list of Listing etc. to table of contents\n"
packages +="\\newcommand{\\source}[1]{\\vspace{-4pt} \\caption*{\\hfill \\footnotesize{Source: {#1}} } }   % Easily insert sources in images\n"
packages +="\n"
packages +="% --- Code ---\n"
packages +="%\\usepackage{minted} % Includes several programming languages and styles, visit --https://www.ntnu.no/wiki/display/imtsoftware/Code+in+LaTeX-- for more information and examples\n"
packages +="\n"
packages +="% --- Colors ---\n"
packages +="\\usepackage[dvipsnames]{xcolor} % Using colors in LaTeX. This package is placed here as it needs to be imported previous to the flow chart packages to ensure no errors.\n"
packages +="\n"
packages +="% --- Flow Charts ---\n"
packages +="% Keep in mind that you may customize these flow chart settings to fit your own preferences.\n"
packages +="\\usepackage{tikz}\n"
packages +="\\usetikzlibrary{shapes,arrows}\n"
packages +="\\tikzstyle{decision} = [diamond, draw, fill=blue!20, \n"
packages +="    text width=4.5em, text badly centered, node distance=3cm, inner sep=0pt]\n"
packages +="\\tikzstyle{block} = [rectangle, draw, fill=blue!20, \n"
packages +="    text width=5em, text centered, rounded corners, minimum height=4em]\n"
packages +="\\tikzstyle{line} = [draw, -latex']\n"
packages +="\\tikzstyle{cloud} = [draw, ellipse,fill=red!20, node distance=3cm,\n"
packages +="    minimum height=2em]\n"
packages +="\n"
packages +="% --- Front matter ---\n"
packages +="% Front matter is located before the table of contents, e.g. preface (Norsk: forord) and often has separate settings\n"
packages +="\\newcommand{\\frontmatter}{\n"
packages +="    \\pagenumbering{roman}   % Setting page numbering to lower-case roman\n"
packages +="}\n"
packages +="\n"
packages +="% --- General ---\n"
packages +="\\usepackage{import}     % Enable importing of sections\n"
packages +="\\usepackage{csquotes}   % Provides international handling of quote marks. Especially useful for bibliography management using BibLaTeX\n"
packages +="\\usepackage{hyperref}   % Hyper-references, possible to change color\n"
packages +="\\hypersetup{    % Color of hyper-references\n"
packages +="    colorlinks,\n"
packages +="    citecolor = black,\n"
packages +="    filecolor = black,\n"
packages +="    linkcolor = black,\n"
packages +="    urlcolor = black\n"
packages +="}\n"
packages +="\\usepackage{comment}    % Comment blocks of text using \\begin{comment} ... \\end{comment}\n"
packages +="\\usepackage{pdfpages}   % Enables import of pdf-pages using e.g. \\includepdf[pages=-]{./my_pdf.pdf}\n"
packages +="\n"
packages +="% --- Image ---\n"
packages +="\\usepackage{graphicx}   % Handle images\n"
packages +="\\usepackage{wrapfig}    % Wrap text around images\n"
packages +="\\usepackage{float}      % Force image location using H\n"
packages +="\\usepackage{url}        % Insert urls\n"
packages +="\\urlstyle{sf}           % Set url-style as sans-serif. Other options are e.g. same or rm\n"
packages +="\\newcommand{\\HRule}{\\rule{\\linewidth}{0.5mm}}   % Ruler\n"
packages +="\n"
packages +="% --- Language ---\n"
packages +="\\usepackage{siunitx}    % Enable SI units\n"
packages +="\\usepackage[portuguese]{babel}     % Defining UK English as language. This will among other things ensure that dates are displayed as 24/03/1997 rather than 03/24/1997 in the bibliography.\n"
packages +="\\addto\\extrasbritish{   % Change naming of different functions, e.g. figure references.\n"
packages +="    \\renewcommand*\\contentsname{Table of Contents}  % Rename table of contents\n"
packages +="    \\renewcommand{\\listfigurename}{List of Figures} % Rename list of figures\n"
packages +="    \\renewcommand{\\listtablename}{List of Tables}   % Rename list of tables\n"
packages +="    \\def\\equationautorefname{Equation}              % Autoref-name for equations\n"
packages +="    \\def\\figureautorefname{Figure}                  % Autoref-name for figures\n"
packages +="    \\def\\tableautorefname{Table}                    % Autoref-name for tables\n"
packages +="    \\def\\sectionautorefname{Section}                % Autoref-name for sections\n"
packages +="    \\def\\subsectionautorefname{\\sectionautorefname} % Autoref-name for subsections\n"
packages +="    \\def\\subsubsectionautorefname{\\sectionautorefname} % Autoref-name for subsubsections\n"
packages +="}\n"
packages +="\n"
packages +="% --- Main matter ---\n"
packages +="% This is the main part of the paper.\n"
packages +="\\newcommand{\\mainmatter}{\n"
packages +="    \\newpage\n"
packages +="    \\pagenumbering{arabic}  % Setting page numbering to normal integers\n"
packages +="}\n"
packages +="\n"
packages +="% --- Mathematics ---\n"
packages +="\\usepackage{bm}         % Bold text in math mode\n"
packages +="\\usepackage{amsmath}    % Math formulas and improved typographical quality of their output\n"
packages +="\\usepackage{amssymb}    % Extended symbol collection\n"
packages +="\\usepackage{amsthm}     % Helps define theorem-like structures\n"
packages +="\\usepackage{textcomp}   % Used in the package gensymb (below), which will give warnings if textcomp is not imported in advance\n"
packages +="\\usepackage{gensymb}    % Adds extra generic symbols for math and text mode, e.g. \\degree\n"
packages +="\n"
packages +="\n"
packages +="% --- Page setup ---\n"
packages +="%\\usepackage[a4paper, total={150mm, 245mm,footskip = 12mm}]{geometry}\n"
packages +="\\setlength{\\parindent}{0.1em}\n"
packages +="\\setlength{\\parskip}{0.8em}\n"
packages +="\n"
packages +="% Customized header and footer\n"
packages +="\\usepackage{fancyhdr}\n"
packages +="\\pagestyle{fancy}\n"
packages +="\\fancyhf{}\n"
packages +="\\renewcommand{\\headrulewidth}{0.1ex}\n"
packages +="\\renewcommand{\\footrulewidth}{0.1ex}\n"
packages +="\\fancyfoot[C]{\\thepage}\n"

# Escrita Packages
with open('packages.sty', mode="w", encoding="utf-8") as h:
    h.write(packages)
